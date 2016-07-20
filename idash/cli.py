import click
import openpyxl
from openpyxl.styles import Alignment, Color, Font, fills, PatternFill
from openpyxl.utils import get_column_letter
import sys
from time import sleep
import collections
from termcolor import colored, cprint
import requests
import datetime
import dateutil.parser
from tqdm import tqdm 

error = lambda x: cprint(x, 'red')
info = lambda x: cprint(x, 'blue')
success = lambda x: cprint(x, 'green')

#Configuration variables
IMPLEMENTATIONS_TEAM_ID = '11068076643777'
PROJECTS_PATH = 'https://app.asana.com/api/1.0/teams/:team-id/projects'
TASKS_PATH = 'https://app.asana.com/api/1.0/projects/:project-id/tasks'
TASKS_OPTIONS = '?opt_fields=completed_at,due_on,name,subtasks,memberships.section.name,id,completed'
TEMPLATE_SHEET_FN = 'template.xlsx'
PROJECT_TEMPLATE_TAB = 'Project Template'
OUTPUT_SHEET_FN = 'Implementation_Dashboard.xlsx'
TEMPLATE_NAME = '[TEMPLATE] Outpatient Services'
MAX_PROJECT_NAME = 31
STRICT_MODE = False
DEFAULT_START_DATE = '01/01/2016'
DEFAULT_GO_LIVE_DATE = '08/01/2016'
SUMMARY_SHEETS = [
  { 'SHEET_NAME': 'Current Imp. Site Stages', 
    'PROJ_ROW_START': 1, 
    'PROJ_COL_START': 3, 
    'TEMPLATE_ROW_START': 2, 
    'TEMPLATE_COL_START': 'B', 
    'TEMPLATE_INCLUDE_TASKS': False, 
    'UPDATE_FORMULAS': False, 
    'WRITE_PROJECTS': True
  },
  { 'SHEET_NAME': 'Implementation Timing',  
    'PROJ_ROW_START': None, 
    'PROJ_COL_START': None, 
    'TEMPLATE_ROW_START': 2, 
    'TEMPLATE_COL_START': 'B',
    'TEMPLATE_INCLUDE_TASKS': True, 
    'UPDATE_FORMULAS': False,
    'WRITE_PROJECTS': False
  },
  { 'SHEET_NAME': 'Task Time v. Start', 
    'PROJ_ROW_START': 3, 
    'PROJ_COL_START': 4, 
    'TEMPLATE_ROW_START': 4, 
    'TEMPLATE_COL_START': 'B',
    'UPDATE_FORMULAS': True,
    'UPDATE_FORMULAS_COL': 'G',
    'TEMPLATE_INCLUDE_TASKS': True, 
    'WRITE_PROJECTS': True
  },
  { 'SHEET_NAME': 'Task Time v. Due', 
    'PROJ_ROW_START': 3, 
    'PROJ_COL_START': 4, 
    'TEMPLATE_ROW_START': 4, 
    'TEMPLATE_COL_START': 'B',
    'UPDATE_FORMULAS': True,
    'UPDATE_FORMULAS_COL': 'H',
    'TEMPLATE_INCLUDE_TASKS': True, 
    'WRITE_PROJECTS': True
  }
]
BOLD_FONT = Font(bold=True, name='Arial', size=11)
REG_FONT = Font(bold=False, name='Arial', size=11)
GRAY_FONT = Font(bold=False, name='Arial', size=11, color='EDEDED')
RED_FONT = Font(bold=False, name='Arial', size=11, color='FF0000')

#Global variables
template = None
template_tasks = None
template_tasks_obj = None
wb = None
token = None
projects = None
filtered_projects = []
project_template_data = {}

@click.command()
@click.option('--write', '-w', help='Specify a filename for the output spreadsheet as "Filename.xlsx"')
@click.option('--template', '-t', help='Specify the name of the Template Task List to retrieve from Asana as "[TEMPLATE] Name"')
@click.option('--strict', '-s', is_flag=True, help='Use strict matching against template task names instead of matching by prefixed task numbers [xx]')
@click.argument('token_file', required=True)
def main(write, template, strict, token_file):
    """Generate the Implementation Dashboard from projects and tasks in Asana"""
    global OUTPUT_SHEET_FN, TEMPLATE_NAME, STRICT_MODE
    if write: 
        if not write.endswith('xlsx'):
            error('Please specify an output file with the extension .xlsx')
            sys.exit(1)
        OUTPUT_SHEET_FN = write
    if template: TEMPLATE_NAME = template
    if strict: STRICT_MODE = True
    readTokenFile(token_file)
    readTemplateSheet() 
    getProjects()
    findTaskTemplate()
    filterTemplates()
    writeProjectData()
    filterProjectsByTemplate()
    writeSummarySheets()
    saveFinalSheet()

def readTokenFile(token_file):
    global token
    try:
        f = open(token_file, 'r')
        token = f.readline()
    except:
        error('Unable to open token file. Please check the path and include the file extension.')
        sys.exit(1)
    if not token:
        error('Unable to read token from file.')
        sys.exit(1)

def readTemplateSheet():
	global wb
	try:
        wb = openpyxl.load_workbook(TEMPLATE_SHEET_FN)
	except:
        error('Unable to read template spreadsheet')
        sys.exit(1)

def getProjects():
    global projects
    token_string = 'Bearer ' + token 
    headers = {'Authorization': token_string}
    path = PROJECTS_PATH.replace(':team-id', IMPLEMENTATIONS_TEAM_ID)
    r = requests.get(path, headers=headers)
    response = r.json()
    if r.status_code != requests.codes.ok:
        error('Error retrieving projects from Asana.')
        error(`r.status_code` + ' ' + response['errors'][0]['message'])
        sys.exit(1)
    projects = response['data']

def writeProjectData():
    global projects
    info('Writing Project Sheets...')
    with tqdm(total=len(projects)) as progress_bar:
        for project in projects:
            tasks = getProjectTasks(project['id'])
            tasks = restructureTaskData(tasks)
            project['tasks'] = tasks
            project_template_data[project['name']] = {}
            writeProjectSheet(project, tasks)
            progress_bar.update(1)

def findTaskTemplate():
    global template, template_tasks, template_tasks_obj
    template = next((project for project in projects if project['name'] == 
      TEMPLATE_NAME), None)
    if not template:
        error("Could not find template '"+ TEMPLATE_NAME +"' from Asana")
        sys.exit(1)
    template_tasks = getProjectTasks(template['id'])
    template_tasks_obj = restructureTaskData(template_tasks)
    if '[0] Metadata:' in template_tasks_obj: del template_tasks_obj['[0] Metadata:']
    project_template_data['template'] = {}

def filterTemplates():
    global projects
    projects = [project for project in projects if (not project['name'].startswith('[TEMPLATE]') and
      not project['name'].startswith('Relicensing'))]

def getProjectTasks(project_id):
    token_string = 'Bearer ' + token 
    headers = {'Authorization': token_string}
    path = TASKS_PATH.replace(':project-id', str(project_id)) + TASKS_OPTIONS
    r = requests.get(path, headers=headers)
    response = r.json()
    if r.status_code != requests.codes.ok:
        error('Error retrieving project tasks from Asana.')
        error(`r.status_code` + ' ' + response['errors'][0]['message'])
        sys.exit(1)
    tasks = response['data']
    return tasks

def restructureTaskData(tasks):
    taskData = collections.OrderedDict()
    for task in tasks:
        if (task['memberships'][0]['section']):
            section = task['memberships'][0]['section']['name']
            if (task['name'] == section):
                taskData[section] = []
            else:
                taskData[section].append(task)
    return taskData

def writeProjectSheet(project, tasks):
    global section_range
    template_sheet = wb.get_sheet_by_name(PROJECT_TEMPLATE_TAB)
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = editProjectName(project['name'])
    start_date = findTaskString(tasks, '[0] Metadata:', 'name', '[0A] ')
    go_live_date = findTaskString(tasks, '[0] Metadata:', 'name', '[0B] ')
    sheet['B1'] = start_date if start_date else DEFAULT_START_DATE
    if not start_date: sheet['B1'].font = RED_FONT 
    sheet['B2'] = go_live_date if go_live_date else DEFAULT_GO_LIVE_DATE
    if not go_live_date: sheet['B2'].font = RED_FONT
    sheet['C1'] = project['name']
    if '[0] Metadata:' in tasks: del tasks['[0] Metadata:']
    row = 2
    project_formulas = {}
    for section in tasks:
        row += 1
        sheet['C'+`row`] = section
        sheet['C'+`row`].font = BOLD_FONT
        startRow = row + 1
        for task in tasks[section]:
            row += 1
            sheet['D'+`row`] = task['name']
            sheet['D'+`row`].font = REG_FONT
            actual_date = findTaskString(task, 'subtasks', 'name', 'Actual Completed Date: ')
            if actual_date: 
                date = dateutil.parser.parse(actual_date)
                sheet['E'+`row`] = date.strftime("%Y-%m-%d")
                sheet['E'+`row`].alignment = Alignment(horizontal='right')
                sheet['E'+`row`].font = REG_FONT
            elif 'completed_at' in task and task['completed_at']: 
                date = dateutil.parser.parse(task['completed_at'])
                sheet['E'+`row`] = date.strftime("%Y-%m-%d")
                sheet['E'+`row`].alignment = Alignment(horizontal='right')
                sheet['E'+`row`].font = REG_FONT
            if 'due_on' in task and task['due_on']:
                date = dateutil.parser.parse(task['due_on'])
                sheet['F'+`row`] = date.strftime("%Y-%m-%d")
                sheet['F'+`row`].alignment = Alignment(horizontal='right')
                sheet['F'+`row`].font = REG_FONT
            project_template_data[project['name']][task['name']] = {'row': row }
        formula = "=countA(EX:EY)/countA(DX:DY)".replace('X', `startRow`).replace('Y', `row`)
        project_formulas[section] = {'formula': formula if tasks[section] else 0 }
    summaryRow = 3
    for section in template_tasks_obj:
        if STRICT_MODE: 
            if section in project_formulas:
                sheet['A'+`summaryRow`] = project_formulas[section]['formula']             
            else:
                sheet['A'+`summaryRow`] = 0 
        else:   
            task_prefix = section[0:section.find(']')+1]
            project_section = next((x for x in project_formulas if x.startswith(task_prefix)), None)
            if project_section:
                sheet['A'+`summaryRow`] = project_formulas[project_section]['formula']
            else:
                sheet['A'+`summaryRow`] = 0  
        sheet['A'+`summaryRow`].font = GRAY_FONT      
        summaryRow += 1

def filterProjectsByTemplate():
    global filtered_projects
    if STRICT_MODE:
        template_names = [task['name'] for task in template_tasks] 
    else:
        template_names = []
        for task in template_tasks:
            if ']' in task['name']:
                task_prefix = task['name'][0:task['name'].find(']')+1]
                template_names.append(task_prefix)
    for project in projects:
        for task in project['tasks']:
            if STRICT_MODE:
                if task in template_names: 
                    filtered_projects.append(project)
                    break
            else:
                if ']' in task:
                    task_prefix = task[0:task.find(']')+1] 
                    if task_prefix in template_names:
                        filtered_projects.append(project)
                        break

def writeSummarySheets():
    info('Writing Summary Sheets...')
    with tqdm(total=len(SUMMARY_SHEETS)) as progress_bar:
        for sheet in SUMMARY_SHEETS:
            writeSheet(sheet)
            sleep(0.1)
            progress_bar.update(1)

def writeSheet(sheet):
    ws = wb.get_sheet_by_name(sheet['SHEET_NAME'])
    if sheet['WRITE_PROJECTS']: writeProjectNames(ws, sheet['PROJ_COL_START'], sheet['PROJ_ROW_START'])
    writeTemplateTasks(ws, sheet['TEMPLATE_ROW_START'], sheet['TEMPLATE_COL_START'], sheet['TEMPLATE_INCLUDE_TASKS'])
    if sheet['UPDATE_FORMULAS']: updateFormulas(ws, sheet['UPDATE_FORMULAS_COL'])

def updateFormulas(sheet, formulaCol):
    for col, project in zip(range(4, len(filtered_projects)+4), filtered_projects):
        row = 3
        for section in template_tasks_obj:
            row += 1
            r = project_template_data['template'][section]
            formula = '=IF(MAX(C@:C#),MAX(C@:C#),"")'.replace('C', get_column_letter(col)).replace('@', 
              `r['start']`).replace('#', `r['end']`)
            sheet.cell(column=col, row=row, value=formula).font = BOLD_FONT
            for task in template_tasks_obj[section]:
                row += 1
                if STRICT_MODE:
                    if task['name'] in project_template_data[project['name']]:
                        cell = formulaCol + `project_template_data[project['name']][task['name']]['row']`
                        formula = (sheet.cell(column=col, row=row).value).replace('XX', cell)
                        sheet.cell(column=col, row=row, value=formula).font = REG_FONT
                else:
                    task_prefix = task['name'][0:task['name'].find(']')+1]
                    matched_task = next((t for t in project_template_data[project['name']] if t.startswith(task_prefix)), None)
                    if matched_task:
                        cell = formulaCol + `project_template_data[project['name']][matched_task]['row']`
                        formula = (sheet.cell(column=col, row=row).value).replace('XX', cell)
                        sheet.cell(column=col, row=row, value=formula).font = REG_FONT

def writeProjectNames(sheet, colStart, row):
    background_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for col, project in zip(range(colStart, len(filtered_projects)+colStart), filtered_projects):
      sheet.cell(column=col, row=row, value=editProjectName(project['name'])).font = BOLD_FONT
      sheet.cell(column=col, row=row).fill = background_fill
      sheet.column_dimensions[get_column_letter(col)].width = MAX_PROJECT_NAME

def writeTemplateTasks(sheet, row, col, include_tasks=True):
    for section in template_tasks_obj:
        sheet[col+`row`] = section
        sheet[col+`row`].font = BOLD_FONT
        startRow = row + 1
        row += 1
        if include_tasks:
            for task in template_tasks_obj[section]:
                sheet[col+`row`] = task['name']
                sheet[col+`row`].font = REG_FONT
                row += 1
        endRow = row if row == startRow else row-1
        project_template_data['template'][section] = {'start': startRow, 'end': endRow}

def editProjectName(name):
    # Remove the prefix in brackets, remove illegal characters, and
    # limit project name to max # of characters as per xlsx rules for tabs
    if name.startswith('[') and name.endswith(']'): name = name[1:-1]
    if ']' in name: name = name.split("]",1)[1].lstrip()
    if '/' in name: name = name.replace('/', '-')
    if (len(name) > MAX_PROJECT_NAME): name = name[0:MAX_PROJECT_NAME]
    return name

def findTaskString(tasks, section, task_attr, prefix):
    # Returns the string found after a given prefix in a task attribute, otherwise returns None
    if section not in tasks: return None
    task = next((t for t in tasks[section] if (task_attr in t and t[task_attr].startswith(prefix))), None)
    if task: return task[task_attr].split(prefix)[1]
    return None

def saveFinalSheet():
    wb.remove_sheet(wb.get_sheet_by_name(PROJECT_TEMPLATE_TAB))
    info('Saving Spreadsheet...')
    wb.save(OUTPUT_SHEET_FN)
    success('Successfully saved ' + OUTPUT_SHEET_FN + '!')
